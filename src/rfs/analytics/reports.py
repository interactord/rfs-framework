"""
RFS Advanced Analytics - Report Generation System (RFS v4.1)

리포트 생성 및 내보내기 시스템
"""

from abc import ABC, abstractmethod
from typing import Any, Dict, List, Optional, Union, BinaryIO
from dataclasses import dataclass, field
from enum import Enum
from datetime import datetime, timedelta
import asyncio
import json
import base64
import io
from pathlib import Path
import tempfile
import uuid

from ..core.types import Result, Success, Failure
from .dashboard import Dashboard
from .charts import Chart
from .data_source import DataSource, DataQuery


class ReportFormat(Enum):
    PDF = "pdf"
    HTML = "html"
    EXCEL = "excel"
    JSON = "json"
    CSV = "csv"


class ReportSchedule(Enum):
    ONCE = "once"
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    YEARLY = "yearly"


@dataclass
class ReportSection:
    """리포트 섹션"""
    section_id: str
    title: str
    content_type: str  # text, chart, table, dashboard, image
    content: Any
    metadata: Dict[str, Any] = field(default_factory=dict)
    
    # 레이아웃 설정
    width: Optional[str] = None  # 100%, 50%, auto
    height: Optional[str] = None
    padding: str = "10px"
    margin: str = "5px"
    
    # 조건부 표시
    condition: Optional[str] = None  # 조건식
    visible: bool = True


@dataclass
class ReportTemplate:
    """리포트 템플릿"""
    template_id: str
    name: str
    description: str
    sections: List[ReportSection]
    variables: Dict[str, Any] = field(default_factory=dict)
    
    # 스타일링
    styles: Dict[str, Any] = field(default_factory=dict)
    header_template: Optional[str] = None
    footer_template: Optional[str] = None
    
    # 메타데이터
    created_at: datetime = field(default_factory=datetime.now)
    updated_at: datetime = field(default_factory=datetime.now)
    version: str = "1.0"


@dataclass
class ReportConfig:
    """리포트 생성 설정"""
    title: str
    subtitle: str = ""
    author: str = ""
    organization: str = ""
    
    # 페이지 설정
    page_size: str = "A4"  # A4, Letter, Legal
    orientation: str = "portrait"  # portrait, landscape
    margin_top: str = "2cm"
    margin_bottom: str = "2cm"
    margin_left: str = "2cm"
    margin_right: str = "2cm"
    
    # 스타일
    font_family: str = "Arial, sans-serif"
    font_size: str = "12pt"
    line_height: str = "1.5"
    
    # 헤더/푸터
    include_header: bool = True
    include_footer: bool = True
    include_page_numbers: bool = True
    include_date: bool = True
    
    # 생성 옵션
    include_toc: bool = False  # Table of Contents
    include_summary: bool = False
    compress: bool = False


class Report(ABC):
    """리포트 추상 클래스"""
    
    def __init__(self, report_id: str, template: ReportTemplate, config: ReportConfig):
        self.report_id = report_id
        self.template = template
        self.config = config
        self.sections: List[ReportSection] = []
        self.variables: Dict[str, Any] = {}
        self.generated_at: Optional[datetime] = None
    
    @abstractmethod
    async def generate(self) -> Result[bytes, str]:
        """리포트 생성"""
        pass
    
    @abstractmethod
    def get_mime_type(self) -> str:
        """MIME 타입 반환"""
        pass
    
    def add_section(self, section: ReportSection) -> Result[bool, str]:
        """섹션 추가"""
        try:
            self.sections.append(section)
            return Success(True)
        except Exception as e:
            return Failure(f"Failed to add section: {str(e)}")
    
    def remove_section(self, section_id: str) -> Result[bool, str]:
        """섹션 제거"""
        try:
            self.sections = [s for s in self.sections if s.section_id != section_id]
            return Success(True)
        except Exception as e:
            return Failure(f"Failed to remove section: {str(e)}")
    
    def set_variable(self, key: str, value: Any) -> Result[bool, str]:
        """변수 설정"""
        try:
            self.variables[key] = value
            return Success(True)
        except Exception as e:
            return Failure(f"Failed to set variable: {str(e)}")
    
    def get_variable(self, key: str, default: Any = None) -> Any:
        """변수 조회"""
        return self.variables.get(key, default)
    
    async def _process_sections(self) -> Result[List[ReportSection], str]:
        """섹션 처리 (조건부 표시, 변수 치환 등)"""
        try:
            processed_sections = []
            
            for section in self.sections:
                # 조건부 표시 처리
                if section.condition:
                    # 간단한 조건 평가 (실제로는 더 복잡한 표현식 파서 필요)
                    if not self._evaluate_condition(section.condition):
                        continue
                
                if not section.visible:
                    continue
                
                # 변수 치환
                processed_section = await self._substitute_variables(section)
                processed_sections.append(processed_section)
            
            return Success(processed_sections)
            
        except Exception as e:
            return Failure(f"Section processing failed: {str(e)}")
    
    def _evaluate_condition(self, condition: str) -> bool:
        """조건 평가 (간단한 구현)"""
        try:
            # 변수를 현재 값으로 치환
            for key, value in self.variables.items():
                condition = condition.replace(f"${{{key}}}", str(value))
            
            # 안전한 평가 (실제로는 더 안전한 파서 사용 권장)
            allowed_names = {"True": True, "False": False, "None": None}
            code = compile(condition, "<string>", "eval")
            return eval(code, {"__builtins__": {}}, allowed_names)
            
        except Exception:
            return True  # 조건 평가 실패시 기본값으로 표시
    
    async def _substitute_variables(self, section: ReportSection) -> ReportSection:
        """변수 치환"""
        # 제목에서 변수 치환
        title = section.title
        for key, value in self.variables.items():
            title = title.replace(f"${{{key}}}", str(value))
        
        # 새로운 섹션 객체 생성 (원본 수정 방지)
        return ReportSection(
            section_id=section.section_id,
            title=title,
            content_type=section.content_type,
            content=section.content,
            metadata=section.metadata,
            width=section.width,
            height=section.height,
            padding=section.padding,
            margin=section.margin,
            condition=section.condition,
            visible=section.visible
        )


class PDFReport(Report):
    """PDF 리포트"""
    
    def get_mime_type(self) -> str:
        return "application/pdf"
    
    async def generate(self) -> Result[bytes, str]:
        """PDF 리포트 생성"""
        try:
            # PDF 생성 라이브러리 사용 (reportlab, weasyprint 등)
            from reportlab.lib.pagesizes import A4, letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
            from reportlab.platypus import Table, TableStyle
            from reportlab.lib import colors
            
            # 임시 파일에 PDF 생성
            buffer = io.BytesIO()
            
            # 페이지 크기 설정
            page_size = A4 if self.config.page_size == "A4" else letter
            
            doc = SimpleDocTemplate(
                buffer,
                pagesize=page_size,
                topMargin=72,
                bottomMargin=72,
                leftMargin=72,
                rightMargin=72
            )
            
            # 스타일 시트
            styles = getSampleStyleSheet()
            
            # 커스텀 스타일 추가
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontSize=24,
                spaceAfter=30,
                alignment=1  # 중앙 정렬
            )
            
            # Story (문서 내용) 생성
            story = []
            
            # 제목
            if self.config.title:
                story.append(Paragraph(self.config.title, title_style))
                story.append(Spacer(1, 20))
            
            # 부제목
            if self.config.subtitle:
                story.append(Paragraph(self.config.subtitle, styles['Heading2']))
                story.append(Spacer(1, 20))
            
            # 메타데이터
            if self.config.author or self.config.organization:
                meta_text = []
                if self.config.author:
                    meta_text.append(f"Author: {self.config.author}")
                if self.config.organization:
                    meta_text.append(f"Organization: {self.config.organization}")
                if self.config.include_date:
                    meta_text.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
                
                story.append(Paragraph("<br/>".join(meta_text), styles['Normal']))
                story.append(Spacer(1, 30))
            
            # 섹션 처리
            sections_result = await self._process_sections()
            if sections_result.is_failure():
                return sections_result
            
            sections = sections_result.unwrap()
            
            for section in sections:
                # 섹션 제목
                if section.title:
                    story.append(Paragraph(section.title, styles['Heading3']))
                    story.append(Spacer(1, 12))
                
                # 섹션 내용
                if section.content_type == "text":
                    story.append(Paragraph(str(section.content), styles['Normal']))
                
                elif section.content_type == "table":
                    if isinstance(section.content, list) and section.content:
                        # 테이블 데이터를 reportlab Table 형식으로 변환
                        if isinstance(section.content[0], dict):
                            # Dict 리스트를 테이블로 변환
                            headers = list(section.content[0].keys())
                            table_data = [headers]
                            for row in section.content:
                                table_data.append([str(row.get(h, "")) for h in headers])
                        else:
                            table_data = section.content
                        
                        table = Table(table_data)
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 14),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        story.append(table)
                
                elif section.content_type == "chart":
                    # 차트를 이미지로 변환하여 삽입
                    if isinstance(section.content, dict) and "image_base64" in section.content:
                        # Base64 이미지 디코딩 및 삽입
                        image_data = base64.b64decode(section.content["image_base64"])
                        img_buffer = io.BytesIO(image_data)
                        
                        from reportlab.lib.utils import ImageReader
                        from reportlab.platypus import Image
                        
                        img = Image(ImageReader(img_buffer), width=400, height=300)
                        story.append(img)
                
                story.append(Spacer(1, 20))
            
            # PDF 생성
            doc.build(story)
            
            self.generated_at = datetime.now()
            
            return Success(buffer.getvalue())
            
        except ImportError:
            return Failure("reportlab library required for PDF generation")
        except Exception as e:
            return Failure(f"PDF generation failed: {str(e)}")


class HTMLReport(Report):
    """HTML 리포트"""
    
    def get_mime_type(self) -> str:
        return "text/html"
    
    async def generate(self) -> Result[bytes, str]:
        """HTML 리포트 생성"""
        try:
            html_parts = []
            
            # HTML 헤더
            html_parts.append(self._generate_html_header())
            
            # 제목 섹션
            html_parts.append(self._generate_title_section())
            
            # 섹션 처리
            sections_result = await self._process_sections()
            if sections_result.is_failure():
                return sections_result
            
            sections = sections_result.unwrap()
            
            # 본문 시작
            html_parts.append('<div class="report-body">')
            
            for section in sections:
                html_parts.append(await self._generate_section_html(section))
            
            html_parts.append('</div>')
            
            # HTML 푸터
            html_parts.append(self._generate_html_footer())
            
            html_content = '\n'.join(html_parts)
            self.generated_at = datetime.now()
            
            return Success(html_content.encode('utf-8'))
            
        except Exception as e:
            return Failure(f"HTML generation failed: {str(e)}")
    
    def _generate_html_header(self) -> str:
        """HTML 헤더 생성"""
        return f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{self.config.title}</title>
    <style>
        body {{
            font-family: {self.config.font_family};
            font-size: {self.config.font_size};
            line-height: {self.config.line_height};
            margin: 0;
            padding: 20px;
            background-color: #f9f9f9;
        }}
        .report-container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            padding: 40px;
            box-shadow: 0 0 10px rgba(0,0,0,0.1);
        }}
        .report-title {{
            text-align: center;
            color: #333;
            border-bottom: 3px solid #007bff;
            padding-bottom: 20px;
            margin-bottom: 30px;
        }}
        .report-subtitle {{
            text-align: center;
            color: #666;
            margin-bottom: 20px;
        }}
        .report-meta {{
            text-align: center;
            color: #888;
            font-size: 0.9em;
            margin-bottom: 40px;
        }}
        .section {{
            margin-bottom: 30px;
        }}
        .section-title {{
            color: #007bff;
            border-left: 4px solid #007bff;
            padding-left: 15px;
            margin-bottom: 15px;
        }}
        .table-container {{
            overflow-x: auto;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
        }}
        th, td {{
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }}
        th {{
            background-color: #007bff;
            color: white;
            font-weight: bold;
        }}
        tr:nth-child(even) {{
            background-color: #f2f2f2;
        }}
        .chart-container {{
            text-align: center;
            margin: 20px 0;
        }}
    </style>
</head>
<body>
<div class="report-container">
"""
    
    def _generate_title_section(self) -> str:
        """제목 섹션 생성"""
        parts = []
        
        if self.config.title:
            parts.append(f'<h1 class="report-title">{self.config.title}</h1>')
        
        if self.config.subtitle:
            parts.append(f'<h2 class="report-subtitle">{self.config.subtitle}</h2>')
        
        meta_parts = []
        if self.config.author:
            meta_parts.append(f"Author: {self.config.author}")
        if self.config.organization:
            meta_parts.append(f"Organization: {self.config.organization}")
        if self.config.include_date:
            meta_parts.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        
        if meta_parts:
            parts.append(f'<div class="report-meta">{" | ".join(meta_parts)}</div>')
        
        return '\n'.join(parts)
    
    async def _generate_section_html(self, section: ReportSection) -> str:
        """섹션 HTML 생성"""
        parts = ['<div class="section">']
        
        if section.title:
            parts.append(f'<h3 class="section-title">{section.title}</h3>')
        
        if section.content_type == "text":
            parts.append(f'<p>{section.content}</p>')
        
        elif section.content_type == "table":
            if isinstance(section.content, list) and section.content:
                parts.append('<div class="table-container">')
                parts.append('<table>')
                
                if isinstance(section.content[0], dict):
                    # Dict 리스트를 테이블로 변환
                    headers = list(section.content[0].keys())
                    parts.append('<thead><tr>')
                    for header in headers:
                        parts.append(f'<th>{header}</th>')
                    parts.append('</tr></thead>')
                    
                    parts.append('<tbody>')
                    for row in section.content:
                        parts.append('<tr>')
                        for header in headers:
                            parts.append(f'<td>{row.get(header, "")}</td>')
                        parts.append('</tr>')
                    parts.append('</tbody>')
                
                parts.append('</table>')
                parts.append('</div>')
        
        elif section.content_type == "chart":
            if isinstance(section.content, dict):
                if "html" in section.content:
                    parts.append('<div class="chart-container">')
                    parts.append(section.content["html"])
                    parts.append('</div>')
                elif "image_base64" in section.content:
                    parts.append('<div class="chart-container">')
                    parts.append(f'<img src="data:image/png;base64,{section.content["image_base64"]}" alt="Chart" style="max-width: 100%; height: auto;">')
                    parts.append('</div>')
        
        parts.append('</div>')
        return '\n'.join(parts)
    
    def _generate_html_footer(self) -> str:
        """HTML 푸터 생성"""
        return """
</div>
</body>
</html>
"""


class ExcelReport(Report):
    """Excel 리포트"""
    
    def get_mime_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    async def generate(self) -> Result[bytes, str]:
        """Excel 리포트 생성"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report"
            
            current_row = 1
            
            # 제목
            if self.config.title:
                ws.cell(row=current_row, column=1, value=self.config.title)
                title_cell = ws.cell(row=current_row, column=1)
                title_cell.font = Font(size=16, bold=True)
                title_cell.alignment = Alignment(horizontal='center')
                current_row += 2
            
            # 메타데이터
            if self.config.author:
                ws.cell(row=current_row, column=1, value=f"Author: {self.config.author}")
                current_row += 1
            
            if self.config.include_date:
                ws.cell(row=current_row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
                current_row += 2
            
            # 섹션 처리
            sections_result = await self._process_sections()
            if sections_result.is_failure():
                return sections_result
            
            sections = sections_result.unwrap()
            
            for section in sections:
                # 섹션 제목
                if section.title:
                    ws.cell(row=current_row, column=1, value=section.title)
                    section_title_cell = ws.cell(row=current_row, column=1)
                    section_title_cell.font = Font(size=14, bold=True)
                    current_row += 1
                
                # 섹션 내용
                if section.content_type == "text":
                    ws.cell(row=current_row, column=1, value=str(section.content))
                    current_row += 1
                
                elif section.content_type == "table":
                    if isinstance(section.content, list) and section.content:
                        if isinstance(section.content[0], dict):
                            # Dict 리스트를 테이블로 변환
                            headers = list(section.content[0].keys())
                            
                            # 헤더 추가
                            for col, header in enumerate(headers, 1):
                                cell = ws.cell(row=current_row, column=col, value=header)
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            current_row += 1
                            
                            # 데이터 추가
                            for row_data in section.content:
                                for col, header in enumerate(headers, 1):
                                    ws.cell(row=current_row, column=col, value=str(row_data.get(header, "")))
                                current_row += 1
                
                current_row += 1  # 섹션 간 공백
            
            # Excel 파일을 바이트로 변환
            buffer = io.BytesIO()
            wb.save(buffer)
            
            self.generated_at = datetime.now()
            
            return Success(buffer.getvalue())
            
        except ImportError:
            return Failure("openpyxl library required for Excel generation")
        except Exception as e:
            return Failure(f"Excel generation failed: {str(e)}")


class ReportBuilder:
    """리포트 빌더"""
    
    def __init__(self):
        self._template: Optional[ReportTemplate] = None
        self._config: Optional[ReportConfig] = None
        self._format: ReportFormat = ReportFormat.HTML
        self._data_sources: Dict[str, DataSource] = {}
    
    def template(self, template: ReportTemplate) -> 'ReportBuilder':
        """템플릿 설정"""
        self._template = template
        return self
    
    def config(self, config: ReportConfig) -> 'ReportBuilder':
        """설정 지정"""
        self._config = config
        return self
    
    def format(self, format: ReportFormat) -> 'ReportBuilder':
        """포맷 설정"""
        self._format = format
        return self
    
    def data_source(self, name: str, source: DataSource) -> 'ReportBuilder':
        """데이터 소스 추가"""
        self._data_sources[name] = source
        return self
    
    async def build(self) -> Result[Report, str]:
        """리포트 생성"""
        if not self._template:
            return Failure("Template not specified")
        
        if not self._config:
            return Failure("Config not specified")
        
        report_id = str(uuid.uuid4())
        
        try:
            if self._format == ReportFormat.PDF:
                report = PDFReport(report_id, self._template, self._config)
            elif self._format == ReportFormat.HTML:
                report = HTMLReport(report_id, self._template, self._config)
            elif self._format == ReportFormat.EXCEL:
                report = ExcelReport(report_id, self._template, self._config)
            else:
                return Failure(f"Unsupported format: {self._format}")
            
            # 템플릿 섹션을 리포트에 복사
            for template_section in self._template.sections:
                report.add_section(template_section)
            
            # 템플릿 변수를 리포트에 복사
            for key, value in self._template.variables.items():
                report.set_variable(key, value)
            
            return Success(report)
            
        except Exception as e:
            return Failure(f"Report build failed: {str(e)}")


async def generate_report(
    template: ReportTemplate,
    config: ReportConfig,
    format: ReportFormat = ReportFormat.HTML,
    variables: Optional[Dict[str, Any]] = None
) -> Result[bytes, str]:
    """리포트 생성 헬퍼 함수"""
    builder = ReportBuilder()
    builder.template(template).config(config).format(format)
    
    report_result = await builder.build()
    if report_result.is_failure():
        return report_result
    
    report = report_result.unwrap()
    
    # 변수 설정
    if variables:
        for key, value in variables.items():
            report.set_variable(key, value)
    
    return await report.generate()


async def schedule_report(
    template: ReportTemplate,
    config: ReportConfig,
    schedule: ReportSchedule,
    format: ReportFormat = ReportFormat.HTML,
    output_path: Optional[str] = None
) -> Result[str, str]:
    """리포트 스케줄링 (간단한 구현)"""
    try:
        # 실제 구현에서는 크론 작업이나 스케줄러 사용
        schedule_id = str(uuid.uuid4())
        
        # 스케줄 정보 저장 (실제로는 데이터베이스 등에 저장)
        schedule_info = {
            "schedule_id": schedule_id,
            "template": template,
            "config": config,
            "schedule": schedule,
            "format": format,
            "output_path": output_path,
            "created_at": datetime.now().isoformat(),
            "next_run": _calculate_next_run(schedule).isoformat()
        }
        
        # 임시로 파일에 저장
        schedule_path = Path(tempfile.gettempdir()) / f"rfs_report_schedule_{schedule_id}.json"
        with open(schedule_path, 'w') as f:
            # JSON 직렬화를 위해 일부 필드 제외
            serializable_info = {
                "schedule_id": schedule_id,
                "template_id": template.template_id,
                "schedule": schedule.value,
                "format": format.value,
                "output_path": output_path,
                "created_at": schedule_info["created_at"],
                "next_run": schedule_info["next_run"]
            }
            json.dump(serializable_info, f, indent=2)
        
        return Success(schedule_id)
        
    except Exception as e:
        return Failure(f"Schedule creation failed: {str(e)}")


def _calculate_next_run(schedule: ReportSchedule) -> datetime:
    """다음 실행 시간 계산"""
    now = datetime.now()
    
    if schedule == ReportSchedule.ONCE:
        return now
    elif schedule == ReportSchedule.DAILY:
        return now + timedelta(days=1)
    elif schedule == ReportSchedule.WEEKLY:
        return now + timedelta(weeks=1)
    elif schedule == ReportSchedule.MONTHLY:
        # 다음 달 같은 날
        if now.month == 12:
            return now.replace(year=now.year + 1, month=1)
        else:
            return now.replace(month=now.month + 1)
    elif schedule == ReportSchedule.QUARTERLY:
        return now + timedelta(days=90)  # 근사치
    elif schedule == ReportSchedule.YEARLY:
        return now.replace(year=now.year + 1)
    
    return now


def create_report_template(
    template_id: str,
    name: str,
    description: str = ""
) -> ReportTemplate:
    """리포트 템플릿 생성 헬퍼 함수"""
    return ReportTemplate(
        template_id=template_id,
        name=name,
        description=description,
        sections=[]
    )


def create_text_section(section_id: str, title: str, content: str) -> ReportSection:
    """텍스트 섹션 생성"""
    return ReportSection(
        section_id=section_id,
        title=title,
        content_type="text",
        content=content
    )


def create_table_section(section_id: str, title: str, data: List[Dict[str, Any]]) -> ReportSection:
    """테이블 섹션 생성"""
    return ReportSection(
        section_id=section_id,
        title=title,
        content_type="table",
        content=data
    )


def create_chart_section(section_id: str, title: str, chart_data: Dict[str, Any]) -> ReportSection:
    """차트 섹션 생성"""
    return ReportSection(
        section_id=section_id,
        title=title,
        content_type="chart",
        content=chart_data
    )